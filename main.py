import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import tkinter as tk
from tkinter import filedialog
import os

# Crear una ventana oculta para usar el cuadro de diálogo de archivo
root = tk.Tk()
root.withdraw()

# Abrir el cuadro de diálogo para que el usuario seleccione el archivo
file_path = filedialog.askopenfilename(
    title="Selecciona el archivo Excel",
    filetypes=[("Archivos Excel", "*.xlsx")]
)

if not file_path:
    print("No se seleccionó ningún archivo. El programa finalizará.")
    exit()

# Cargar todo el contenido del archivo
df = pd.read_excel(file_path, sheet_name='Rpt210_ReporteLiquidaciones', header=None)

# Definir el encabezado general que se desea conservar
encabezado_referencia = [
    "Liquidación", "Fecha liquidación", "Automática", "Estado", "Oracle ID", "Número viaje", 
    "Número OT", "Peso entregado", "Fecha Creación Viaje", "Manifiesto", "Fecha creación Manifiesto", 
    "Ciudad origen", "Ciudad destino", "Código poseedor", "Poseedor", "Sucursal", 
    "Código conductor", "Conductor", "Tipo operación", "Placa", "Remolque", "Precio bruto", 
    "Total anticipos", "Descuento", "Precio neto", "Usuario creación", "Fecha creación", 
    "Usuario modificación", "Fecha modificación"
]

# Crear una función para identificar filas a eliminar basadas en el patrón "Cédula:" y "Poseedor:"
def es_fila_innecesaria(fila):
    fila_sin_nan = [str(item).strip() for item in fila]
    return any("Cédula:" in item for item in fila_sin_nan) or any("Poseedor:" in item for item in fila_sin_nan)

# Filtrar las filas que no son innecesarias
df_filtrado = df[~df.apply(es_fila_innecesaria, axis=1)]

# Eliminar las primeras 12 filas
df_filtrado = df_filtrado.iloc[12:]

# Eliminar columnas que solo contienen valores vacíos o espacios en blanco
df_filtrado.replace(r'^\s*$', pd.NA, regex=True, inplace=True)
df_filtrado.dropna(how='all', axis=1, inplace=True)

# Restablecer los índices
df_filtrado.reset_index(drop=True, inplace=True)

# Agregar el encabezado general al inicio
if df_filtrado.shape[1] == len(encabezado_referencia):
    df_filtrado.columns = encabezado_referencia
else:
    print("Error: El número de columnas no coincide con el número de encabezados.")

# Eliminar filas que son idénticas a los encabezados
df_filtrado = df_filtrado[~df_filtrado.isin(encabezado_referencia).all(axis=1)]

# Guardar la tabla filtrada en un archivo Excel
output_file = 'Liquidaciones_organizadas_sin_espacios_y_cedula.xlsx'
df_filtrado.to_excel(output_file, index=False)

# Cargar el archivo para aplicar formato
wb = load_workbook(output_file)
ws = wb.active

# Establecer el color del encabezado
color_header = "EFA779"  

# Aplicar formato al encabezado
for cell in ws[1]:
    cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.alignment = cell.alignment.copy(horizontal='center')  # Centrar el texto en el encabezado

# Aplicar formato de moneda a las columnas específicas
moneda_columnas = ["Precio bruto", "Total anticipos", "Descuento", "Precio neto"]
for col in moneda_columnas:
    col_idx = df_filtrado.columns.get_loc(col) + 1  # +1 porque openpyxl es 1-indexado
    for row in range(2, len(df_filtrado) + 2):  # +2 para saltar el encabezado
        cell = ws.cell(row=row, column=col_idx)
        cell.number_format = '"$"#,##0.00'  # Formato de moneda

# Ajustar el ancho de las columnas
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter  # Obtener la letra de la columna
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)  # Agregar un margen de 2
    ws.column_dimensions[column_letter].width = adjusted_width

# Guardar el archivo con los cambios
wb.save(output_file)

print(f"El archivo ha sido procesado y guardado como '{output_file}'.")
